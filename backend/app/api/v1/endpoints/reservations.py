"""
API Router para la gestión de reservas de restaurante.
Este endpoint maneja el CRUD completo de reservas.
"""

from fastapi import APIRouter, Depends, HTTPException, status, Query
from sqlalchemy.orm import Session
from datetime import datetime, date as date_class, timedelta
from typing import List, Optional
import csv
import io

from app.db.session import get_db
from app.services.reservation_service import ReservationService
from app.schemas.reservations import (
    ReservationCreate, 
    ReservationRead, 
    ReservationUpdate,
    RestaurantAvailabilityResponse
)

router = APIRouter()


@router.post("/", response_model=ReservationRead, status_code=status.HTTP_201_CREATED)
def create_reservation(
    reservation_data: ReservationCreate, 
    db: Session = Depends(get_db)
):
    """
    Crea una nueva reserva de restaurante.
    """
    try:
        service = ReservationService(db)
        return service.create_reservation(reservation_data)
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        print(f"Error creando reserva: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno al crear la reserva"
        )


@router.get("/", response_model=List[ReservationRead])
def get_reservations(
    date: Optional[str] = Query(None, description="Fecha en formato YYYY-MM-DD"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: Session = Depends(get_db)
):
    """
    Obtiene lista de reservas.
    Si se proporciona fecha, filtra por ese día.
    """
    try:
        service = ReservationService(db)
        
        if date:
            target_date = datetime.strptime(date, "%Y-%m-%d").date()
            return service.get_reservations_by_date(target_date)
        else:
            # Si no hay fecha, devolver reservas de hoy en adelante
            today = date_class.today()
            all_reservations = []

            # Obtener reservas de los próximos 30 días
            for i in range(30):
                current_date = today + timedelta(days=i)
                day_reservations = service.get_reservations_by_date(current_date)
                all_reservations.extend(day_reservations)
            
            return all_reservations[skip:skip + limit]
            
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato de fecha inválido. Use YYYY-MM-DD"
        )
    except Exception as e:
        print(f"Error obteniendo reservas: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno al obtener las reservas"
        )


@router.get("/{reservation_id}", response_model=ReservationRead)
def get_reservation(reservation_id: int, db: Session = Depends(get_db)):
    """
    Obtiene una reserva específica por su ID.
    """
    try:
        service = ReservationService(db)
        reservation = service.get_reservation_by_id(reservation_id)
        
        if not reservation:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Reserva no encontrada"
            )
        
        return reservation
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error obteniendo reserva: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno al obtener la reserva"
        )


@router.put("/{reservation_id}", response_model=ReservationRead)
def update_reservation(
    reservation_id: int,
    reservation_update: ReservationUpdate,
    db: Session = Depends(get_db)
):
    """
    Actualiza una reserva existente.
    """
    try:
        service = ReservationService(db)
        updated_reservation = service.update_reservation(reservation_id, reservation_update)
        
        if not updated_reservation:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Reserva no encontrada"
            )
        
        return updated_reservation
        
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error actualizando reserva: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno al actualizar la reserva"
        )


@router.delete("/{reservation_id}")
def delete_reservation(reservation_id: int, db: Session = Depends(get_db)):
    """
    Cancela una reserva (soft delete).
    """
    try:
        service = ReservationService(db)
        success = service.cancel_reservation(reservation_id)
        
        if not success:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Reserva no encontrada"
            )
        
        return {"message": "Reserva cancelada correctamente"}
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error cancelando reserva: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno al cancelar la reserva"
        )


@router.post("/{reservation_id}/confirm")
def confirm_reservation(reservation_id: int, db: Session = Depends(get_db)):
    """
    Confirma una reserva.
    """
    try:
        service = ReservationService(db)
        success = service.confirm_reservation(reservation_id)
        
        if not success:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Reserva no encontrada o no está en estado programado"
            )
        
        return {"message": "Reserva confirmada correctamente"}
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error confirmando reserva: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno al confirmar la reserva"
        )


@router.get("/client/phone/{phone}", response_model=List[ReservationRead])
def get_reservations_by_phone(phone: str, db: Session = Depends(get_db)):
    """
    Obtiene todas las reservas de un cliente por su teléfono.
    """
    try:
        service = ReservationService(db)
        return service.get_reservations_by_phone(phone)
        
    except Exception as e:
        print(f"Error obteniendo reservas por teléfono: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno al obtener las reservas"
        )


@router.get("/daily-summary/{date}")
def get_daily_summary(date: str, db: Session = Depends(get_db)):
    """
    Obtiene un resumen de ocupación para un día específico.
    """
    try:
        target_date = datetime.strptime(date, "%Y-%m-%d").date()
        service = ReservationService(db)
        return service.get_daily_capacity_summary(target_date)
        
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato de fecha inválido. Use YYYY-MM-DD"
        )
    except Exception as e:
        print(f"Error obteniendo resumen diario: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno al obtener el resumen"
        )


# ─── Historial, Estadísticas y Exportación ─────────────────────────────


@router.get("/history")
def get_reservation_history(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Obtiene historial paginado de reservas con filtros."""
    try:
        service = ReservationService(db)

        start = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
        end = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None

        return service.get_history(
            page=page,
            per_page=per_page,
            start_date=start,
            end_date=end,
            status_filter=status,
            search=search,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")
    except Exception as e:
        print(f"Error en historial: {e}")
        raise HTTPException(status_code=500, detail="Error interno al obtener el historial")


@router.get("/stats")
def get_reservation_stats(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Obtiene estadísticas agregadas del historial."""
    try:
        service = ReservationService(db)

        start = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
        end = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None

        return service.get_stats(start_date=start, end_date=end)
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")
    except Exception as e:
        print(f"Error en stats: {e}")
        raise HTTPException(status_code=500, detail="Error interno al obtener estadísticas")


@router.get("/export/csv")
def export_reservations_csv(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Exporta reservas a CSV."""
    try:
        service = ReservationService(db)
        start = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
        end = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None

        result = service.get_history(
            page=1, per_page=10000,
            start_date=start, end_date=end,
            status_filter=status,
        )
        reservations = result["items"]

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["ID", "Fecha", "Hora", "Cliente", "Teléfono", "Email", "Personas", "Estado", "Origen", "Notas", "Creado"])

        for r in reservations:
            start_dt = r.start_time.strftime("%Y-%m-%d %H:%M") if r.start_time else ""
            writer.writerow([
                r.id, start_dt.split(" ")[0] if start_dt else "",
                start_dt.split(" ")[1] if " " in start_dt else "",
                r.client_name, r.client_phone, r.client_email or "",
                r.party_size, r.status, r.source or "manual",
                r.client_notes or "", r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
            ])

        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=historial_reservas.csv"},
        )
    except Exception as e:
        print(f"Error exportando CSV: {e}")
        raise HTTPException(status_code=500, detail="Error al exportar CSV")


@router.get("/export/excel")
def export_reservations_excel(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Exporta reservas a Excel (.xlsx)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        service = ReservationService(db)
        start = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
        end = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None

        result = service.get_history(
            page=1, per_page=10000,
            start_date=start, end_date=end,
            status_filter=status,
        )
        reservations = result["items"]

        wb = Workbook()
        ws = wb.active
        ws.title = "Historial de Reservas"

        # Encabezados
        headers = ["ID", "Fecha", "Hora", "Cliente", "Teléfono", "Email", "Personas", "Estado", "Origen", "Notas", "Creado"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Datos
        for i, r in enumerate(reservations, 2):
            start_dt = r.start_time.strftime("%Y-%m-%d %H:%M") if r.start_time else ""
            ws.cell(row=i, column=1, value=r.id)
            ws.cell(row=i, column=2, value=start_dt.split(" ")[0] if start_dt else "")
            ws.cell(row=i, column=3, value=start_dt.split(" ")[1] if " " in start_dt else "")
            ws.cell(row=i, column=4, value=r.client_name)
            ws.cell(row=i, column=5, value=r.client_phone)
            ws.cell(row=i, column=6, value=r.client_email or "")
            ws.cell(row=i, column=7, value=r.party_size)
            ws.cell(row=i, column=8, value=r.status)
            ws.cell(row=i, column=9, value=r.source or "manual")
            ws.cell(row=i, column=10, value=r.client_notes or "")
            ws.cell(row=i, column=11, value=r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "")

        # Ajustar ancho de columnas
        col_widths = [6, 12, 8, 25, 15, 25, 10, 14, 10, 30, 18]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else ""].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=historial_reservas.xlsx"},
        )
    except Exception as e:
        print(f"Error exportando Excel: {e}")
        raise HTTPException(status_code=500, detail="Error al exportar Excel")
